import os
import json
import pandas as pd
import re
import time

from controllers.invoice_processing.invoice_extractor import InvoiceExtractor
from extract_msg.exceptions import InvalidFileFormatError
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

class InvoiceProcessor:
    def __init__(self, params, translations, window_instance):

        self.config = params
        self.translations = translations
        self.txtEvents = window_instance.txtEvents
        self.tiempo = 0
        self.tiempo_ini = None 

        if ( self.config['email_file_type'] == '.eml' ):
            self.emails_folder = self.config[ 'path_emails_eml' ]
        else:
            self.emails_folder = self.config[ 'path_emails_msg' ]

        self.openai_config = {
            'api_key': self.config['api_key'],
            'type_content': self.config['type_content'],
            'prompts': self.config['prompts'],
            'optional_prompts': self.config['optional_prompts'],
            'minimum_key_count': self.config['minimum_key_count'],
            'max_nrquest_openai': self.config['max_nrquest_openai'],
            'email_init': self.config['email_init'],
        }

        params = {k: v for k, v in params.items() if k not in ['api_key', 'email', 'password']}
        prompts_df = pd.DataFrame(columns=['ID', 'PROMPT']) # Crea un DataFrame vacío


        for prompt in self.openai_config['prompts']: # Recorriendo la lista de prompts
            # Crear un DataFrame temporal para cada nuevo registro
            temp_df = pd.DataFrame([{'ID': prompt['ID'], 'PROMPT': prompt['PROMPT']}])
            # Concatenar este DataFrame con el DataFrame original
            prompts_df = pd.concat([prompts_df, temp_df], ignore_index=True)

        prompts_df.reset_index(drop=True, inplace=True) # Reset del índice del DataFrame

        self.params_df = pd.DataFrame.from_dict(params, orient='index', columns=['Value'])
        self.params_df.reset_index(inplace=True)
        self.params_df.columns = ['Parameter', 'Value']

        # Concatenar el DataFrame de parámetros con el DataFrame de prompts
        self.params_df = pd.concat([self.params_df, prompts_df], axis=0).reset_index(drop=True)

        # Encuentra la fila en la que se almacenará la ruta en el archivo de Excel
        self.path_parameter_excel_row = self.params_df[self.params_df['Value'] == self.emails_folder].index[0] + 2

    def process_invoices_offline(self):
        self.tiempo_ini = time.time()  
        self.txtEvents.clear()  # Limpia los eventos de texto
        invoice_extractor = InvoiceExtractor(self.config, self.translations, self.openai_config, self.txtEvents, self)  # Crea una instancia de InvoiceExtractor
        all_header = []  # Inicializa la lista de todos los encabezados
        all_detail = []  # Inicializa la lista de todos los detalles
        all_headero = []  # Inicializa la lista de todos los encabezados de orden
        all_detailo = []  # Inicializa la lista de todos los detalles de orden
        n_request = 1  # Inicializa el número de solicitud
        start_at_email = self.config.get('email_init', 0)  # Obtiene el inicio del correo electrónico desde la configuración
        max_invoice = self.config.get('max_invoice', -1)  # Obtiene el número máximo de facturas desde la configuración

        files = os.listdir(self.emails_folder)  # Lista todos los archivos en la carpeta de correos electrónicos
        files = [f for f in files if f.endswith('.msg') or f.endswith('.eml')]  # Filtra los archivos que terminan en .msg o .eml

        # Limita el número de archivos procesados en base a max_invoice
        files_to_process = files[start_at_email - 1:start_at_email - 1 + max_invoice]  # Obtiene los archivos a procesar
         
        for i, msg_file in enumerate(files_to_process):  # Itera sobre los archivos a procesar
            try:
                file_path = os.path.join(self.emails_folder, msg_file)  # Obtiene la ruta del archivo
                n_request, headeri_data, detaili_data, headero_data, detailo_data, all_info = invoice_extractor.extract_data_from_file(n_request, file_path)  # Extrae los datos del archivo
                # Convertir el nombre del archivo en un enlace HTML
                file_link = os.path.join(self.emails_folder, msg_file)  # Obtiene el enlace del archivo
                file_link_html = f'=HYPERLINK("{file_link}", "{msg_file}")'  # Crea el enlace HTML del archivo

                for header_item in headeri_data:  # Itera sobre los elementos del encabezado
                    # Añadir el nombre del archivo a los datos transformados
                    header_item['file'] = file_link_html  # Añade el enlace del archivo al elemento del encabezado
                    header_item['n_email'] = i + 1  # Añade el número de correo electrónico al elemento del encabezado
                    transformed_header_data = InvoiceProcessor.transform_header_data(header_item)  # Transforma los datos del encabezado
                    all_header.append(transformed_header_data)  # Añade los datos transformados del encabezado a la lista de todos los encabezados

                for detail_item in detaili_data:  # Itera sobre los elementos de detalle
                    # Añadir el nombre del archivo a los datos transformados
                    detail_item['file'] = file_link_html  # Añade el enlace del archivo al elemento de detalle
                    detail_item['n_email'] = i + 1  # Añade el número de correo electrónico al elemento de detalle
                    transformed_detail_data = InvoiceProcessor.transform_detail_data(detail_item)  # Transforma los datos de detalle
                    all_detail.extend(transformed_detail_data)  # Añade los datos transformados del detalle a la lista de todos los detalles

                for header_item in headero_data:  # Itera sobre los elementos del encabezado de orden
                    # Añadir el nombre del archivo a los datos transformados
                    header_item['file'] = file_link_html  # Añade el enlace del archivo al elemento del encabezado de orden
                    header_item['n_email'] = i + 1  # Añade el número de correo electrónico al elemento del encabezado de orden
                    transformed_headero_data = InvoiceProcessor.transform_header_data(header_item)  # Transforma los datos del encabezado de orden
                    all_headero.append(transformed_headero_data)  # Añade los datos transformados del encabezado de orden a la lista de todos los encabezados de orden

                for detail_item in detailo_data:  # Itera sobre los elementos de detalle de orden
                    # Añadir el nombre del archivo a los datos transformados
                    detail_item['file'] = file_link_html  # Añade el enlace del archivo al elemento de detalle de orden
                    detail_item['n_email'] = i + 1  # Añade el número de correo electrónico al elemento de detalle de orden
                    transformed_detailo_data = InvoiceProcessor.transform_detail_data(detail_item)  # Transforma los datos de detalle de orden
                    all_detailo.extend(transformed_detailo_data)  # Añade los datos transformados del detalle de orden a la lista de todos los detalles de orden
                for info in all_info:  # Itera sobre todos los elementos de información
                    if info.get('n_email') is None:  # Si el elemento de información no tiene un número de correo electrónico
                        info['file'] = file_link_html  # Añade el enlace del archivo al elemento de información
                        info['n_email'] = i+1  # Añade el número de correo electrónico al elemento de información

            except InvalidFileFormatError:  # Si se produce un error de formato de archivo no válido
                self.txtEvents.insertPlainText(f"{self.translations['ui']['err1']} {msg_file}\n")  # Inserta un mensaje de error en los eventos de texto

        # Guardar los resultados en un archivo de Excel
        self._save_results_to_excel(all_header, all_detail, all_headero, all_detailo, all_info)  # Guarda los resultados en un archivo de Excel

    def _save_results_to_excel(self, all_header_data, all_detail_data,all_headero_data, all_detailo_data, all_extra_data):
        # Crear un archivo de Excel con múltiples hojas
        output_file_path = os.path.join(self.config['path_export_xls'], 'result.xlsx')  # Obtiene la ruta del archivo de salida

        # Inicializar DataFrames vacíos para almacenar todos los datos de cabecera, detalles y extra
        all_header_df = pd.DataFrame()  # Inicializa el DataFrame de todos los encabezados
        all_detail_df = pd.DataFrame()  # Inicializa el DataFrame de todos los detalles
        all_headero_df = pd.DataFrame()  # Inicializa el DataFrame de todos los encabezados de orden
        all_detailo_df = pd.DataFrame()  # Inicializa el DataFrame de todos los detalles de orden
        all_extra_df = pd.DataFrame()  # Inicializa el DataFrame de todos los datos extra

        header_order = self.translations['invoice_header_order'] 
        detail_order = self.translations['invoice_detail_order']  # Obtiene el orden de los detalles de la factura desde las traducciones
        headero_order = self.translations['order_header_order']  # Obtiene el orden de los encabezados de orden desde las traducciones
        detailo_order = self.translations['order_detail_order']  # Obtiene el orden de los detalles de orden desde las traducciones

       
        for header_data in all_header_data:  # Iterar sobre cada factura en la cabecera individualmente
            df_header = pd.DataFrame([header_data]) # Crear un DataFrame de pandas con los datos y las columnas
            all_header_df = pd.concat([all_header_df, df_header])  # Añadir los nuevos datos al DataFrame acumulativo de cabecera

        
        for headero_data in all_headero_data: # Iterar sobre cada orden en la cabecera individualmente
            df_headero = pd.DataFrame([headero_data])  # Crear un DataFrame de pandas con los datos y las columnas
            all_headero_df = pd.concat([all_headero_df, df_headero]) # Añadir los nuevos datos al DataFrame acumulativo de cabecera


        for detail_data in all_detail_data:    # Iterar sobre cada detalle individualmente
            df_detail = pd.DataFrame([detail_data]) # Añadir los nuevos datos al DataFrame acumulativo de detalles
            all_detail_df = pd.concat([all_detail_df, df_detail])

        for detailo_data in all_detailo_data: # Iterar sobre cada detalle de una orden individualmente
            df_detailo = pd.DataFrame([detailo_data])
            all_detailo_df = pd.concat([all_detailo_df, df_detailo])  # Añadir los nuevos datos al DataFrame acumulativo de detalles

        if isinstance(all_extra_data, dict):  # Si all_extra_data es un diccionario, convertirlo en una lista de un solo elemento
            all_extra_data = [all_extra_data]

        # Obtener todas las columnas posibles que pueden estar en los diccionarios
        all_keys = set().union(*(d.keys() for d in all_extra_data if isinstance(d, dict)))
        all_keys.update(['n_email', 'file'])  # Asegura que 'n_email' y 'file' estén en all_keys

       
        for extra_data in all_extra_data:  # Iterar sobre cada conjunto de datos extra individualmente
            for key in all_keys:  # Asegurar de que cada diccionario tenga todas las claves posibles
                if key not in extra_data:
                    extra_data[key] = None  # o cualquier otro valor que quieras usar para representar datos faltantes

            df_extra = pd.DataFrame([extra_data])
            all_extra_df = pd.concat([all_extra_df, df_extra]) # Añadir los nuevos datos al DataFrame acumulativo de datos extra

        if not all_extra_df.empty:
                cols_to_include_extra = ['n_email', 'file'] + [col for col in all_extra_df.columns if col not in ['n_email', 'file']]
                all_extra_df = all_extra_df[cols_to_include_extra]

        if not all_header_df.empty:
            if 'n_email' in all_header_df.columns and 'file' in all_header_df.columns:
                cols_to_include_header =['n_email', 'file'] + [col for col in header_order if col in all_header_df.columns]
                all_header_df = all_header_df[cols_to_include_header]

        if not all_headero_df.empty:
            if 'n_email' in all_headero_df.columns and 'file' in all_headero_df.columns:
                cols_to_include_headero = ['n_email', 'file'] + [col for col in headero_order if col in all_headero_df.columns]
                all_headero_df = all_headero_df[cols_to_include_headero]

        if not all_detail_df.empty:
            if 'n_email' in all_detail_df.columns and 'file' in all_detail_df.columns:
                cols_to_include_detail = ['n_email', 'file'] + [col for col in detail_order if col in all_detail_df.columns]
                all_detail_df = all_detail_df[cols_to_include_detail]

        if not all_detailo_df.empty:
            if 'n_email' in all_detailo_df.columns and 'file' in all_detailo_df.columns:
                cols_to_include_detailo = ['n_email', 'file'] + [col for col in detailo_order if col in all_detailo_df.columns]
                all_detailo_df = all_detailo_df[cols_to_include_detailo]

        try:
            with pd.ExcelWriter(output_file_path) as writer: # Escribir los DataFrames acumulativos en Excel
                if not all_header_df.empty:
                    self.clean_df(all_header_df).to_excel(writer, sheet_name='Inv Header Dat', index=False)
                if not all_detail_df.empty:
                    self.clean_df(all_detail_df).to_excel(writer, sheet_name='Inv Detail Dat', index=False)
                if not all_headero_df.empty:
                    self.clean_df(all_headero_df).to_excel(writer, sheet_name='Ord Header Dat', index=False)
                if not all_detailo_df.empty:
                    self.clean_df(all_detailo_df).to_excel(writer, sheet_name='Ord Detail Dat', index=False)
                if not all_extra_df.empty:
                    self.clean_df(all_extra_df).to_excel(writer, sheet_name='Response Data', index=False)

                self.params_df.to_excel(writer, sheet_name='Configuration', index=False)  # Guardar las variables de configuración

        except PermissionError:
            self.txtEvents.insertPlainText(f"Error: No se pudo escribir en el archivo '{output_file_path}'. Verifica que el archivo no esté abierto en otra aplicación y que tengas permisos de escritura.\n")

        else:
            self.adjust_columns_width_and_rows_height(output_file_path)   # Si no hubo errores de permisos, ajusta las columnas y las filas        

    def clean_df(self, df):
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].apply(lambda x: re.sub(r'[^\x20-\x7E]', '', str(x)) if x else x)
        return df

    @staticmethod
    def transform_detail_data(detail_data):
        transformed_data = []
        keys = detail_data.keys()

        # Verificar si hay al menos un elemento que sea una lista antes de calcular el valor máximo
        if any(isinstance(value, list) for value in detail_data.values()):
            max_length = max(len(value) for value in detail_data.values() if isinstance(value, list))
        else:
            # Si no hay listas en los valores, devolver un objeto con los valores de entrada
            return [detail_data]

        for i in range(max_length):
            item = {}
            for key in keys:
                if isinstance(detail_data[key], list):
                    if i < len(detail_data[key]):
                        value = detail_data[key][i]

                        # Redondear valores numéricos a dos decimales
                        if key in ['Quantity', 'Price', 'Total Line','Total Line Calculated', 'Confidence'] and isinstance(value, (int, float)):
                            value = round(value, 2)

                        item[key] = value
                    else:
                        item[key] = None
                else:
                    item[key] = detail_data[key]
            transformed_data.append(item)

        return transformed_data     
   
    @staticmethod
    def transform_header_data(header_data):
        transformed_data = {}
        keys = header_data.keys()

        for key in keys:
            if isinstance(header_data[key], list):
                for i, value in enumerate(header_data[key]):
                    transformed_data[f"{key}_{i + 1}"] = value
            else:
                transformed_data[key] = header_data[key]

        return transformed_data


    @staticmethod
    def load_column_name_mapping(json_filepath):
        with open(json_filepath, "r") as file:
            column_name_mapping = json.load(file)
        return column_name_mapping

    @staticmethod
    def normalize_labels(data, label_mapping):
        if isinstance(data, list):
            return [InvoiceProcessor.normalize_labels(item, label_mapping) for item in data]
        elif isinstance(data, dict):
            normalized_data = {}
            for key, value in data.items():
                normalized_key = None
                for normalized_column, possible_columns in label_mapping.items():
                    if key in possible_columns:
                        normalized_key = normalized_column
                        break
                if normalized_key is not None:
                    normalized_data[normalized_key] = value
                else:
                    normalized_data[key] = value
            return normalized_data
        else:
            raise ValueError("Invalid data type. Expected a dict or list of dicts.")


    def adjust_columns_width_and_rows_height(self, path):
        # Cargar el libro de trabajo
        book = load_workbook(path)
        
        for index, sheet in enumerate(book.worksheets):
            # Iterar sobre las columnas
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                # Encontrar el largo máximo en la columna
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                # Ajustar el ancho de la columna al máximo encontrado (con un mínimo y un máximo)
                adjusted_width = (max_length + 2) if max_length > 0 else 10
                adjusted_width = min(50, max(10, adjusted_width))  # Ajustar al rango 10-50
                sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
                # Ajustar el ancho de las columnas específicas
                sheet.column_dimensions['A'].width = max(len("Email"), 10)
                sheet.column_dimensions['B'].width = max(len("RequestNumber") + 3, 10)
                sheet.column_dimensions['C'].width = max(len("Classify") + 10, 10)
                sheet.column_dimensions['H'].width = max(10, len("VERDADERO") + 4)  

            # Si es la última hoja, ajustar el alto de las filas y el texto
            if index == len(book.worksheets) - 1:
                # Ajustar el alto de las filas en la última hoja
                for row in sheet.iter_rows(min_row = 2):
                    sheet.row_dimensions[row[0].row].height = 15 # Cambiar este valor al que necesites

                # Ajustar el texto en la última hoja
                for row in sheet.iter_rows(min_row = 1, max_row = 1): # Ajustar solo la primera fila (cabecera)
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                for row in sheet.iter_rows(min_row = 2): # Ajustar el resto de las filas
                    for cell in row:
                        # Alinear las tres primeras columnas de manera centrada
                        if cell.column <= 3:
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        elif 3 < cell.column <= 8:
                            cell.alignment = Alignment(vertical="top", wrap_text=True)
                        # Ajustar las celdas en todas las otras columnas solo tienen ajuste de texto
                        else:
                            cell.alignment = Alignment(wrap_text=True)
                    
        # Guardar el libro con los cambios
        book.save(path)